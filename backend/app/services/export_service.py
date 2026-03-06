"""
Exportación del ranking de un proceso a un archivo Excel (.xlsx).
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.logger import get_logger

logger = get_logger(__name__)

# Colores
COLOR_HEADER    = "1F4E79"
COLOR_SUBHEADER = "2E75B6"
COLOR_GREEN     = "E2EFDA"
COLOR_YELLOW    = "FFFCE6"
COLOR_RED       = "FCE4D6"
COLOR_WHITE     = "FFFFFF"
COLOR_GRAY      = "F2F2F2"


def generar_excel_ranking(proceso_nombre: str, items: list, destino: Path) -> Path:
    """
    Genera el archivo Excel con el ranking completo.
    items: lista de dicts con {posicion, candidato, analisis}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Título ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Ranking de Candidatos — {proceso_nombre}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=COLOR_WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Encabezados ───────────────────────────────────────────────────────
    headers = ["#", "Nombre", "Email", "Teléfono", "Puntaje", "Estado IA", "Proveedor IA", "Recomendación IA"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
    ws.row_dimensions[2].height = 22

    # ── Datos ─────────────────────────────────────────────────────────────
    for i, item in enumerate(items, start=3):
        c    = item["candidato"]
        an   = item["analisis"]
        pos  = item["posicion"]
        puntaje = an.puntaje_total if an else None

        # Color de fila según puntaje
        if puntaje is None:
            fill_color = COLOR_GRAY
        elif puntaje >= 70:
            fill_color = COLOR_GREEN
        elif puntaje >= 50:
            fill_color = COLOR_YELLOW
        else:
            fill_color = COLOR_RED

        fila = [
            pos,
            c.nombre or "—",
            c.email or "—",
            c.telefono or "—",
            f"{puntaje:.1f}%" if puntaje is not None else "—",
            an.estado if an else "sin analizar",
            an.proveedor_ia or "—" if an else "—",
            (an.resumen_ia[:200] if an and an.resumen_ia else "—"),
        ]

        for col, val in enumerate(fila, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 8))
            cell.border = bdr

        ws.row_dimensions[i].height = 18

    # ── Anchos de columna ─────────────────────────────────────────────────
    anchos = [5, 28, 30, 16, 10, 14, 14, 60]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # ── Hoja de detalle por candidato ─────────────────────────────────────
    ws2 = wb.create_sheet(title="Detalle Criterios")
    det_headers = ["Candidato", "Criterio", "Cumple", "Puntaje Criterio", "Descripción"]
    for col, h in enumerate(det_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.border = bdr

    fila_det = 2
    for item in items:
        c  = item["candidato"]
        an = item["analisis"]
        if not an or not an.detalle_json:
            continue
        for crit in an.detalle_json:
            ws2.cell(row=fila_det, column=1, value=c.nombre or "—").border = bdr
            ws2.cell(row=fila_det, column=2, value=crit.get("criterio", "")).border = bdr
            ws2.cell(row=fila_det, column=3, value=crit.get("cumple", "")).border = bdr
            ws2.cell(row=fila_det, column=4, value=crit.get("puntaje", "")).border = bdr
            ws2.cell(row=fila_det, column=5, value=crit.get("descripcion", "")).border = bdr
            fila_det += 1

    for col, ancho in enumerate([28, 30, 10, 14, 60], 1):
        ws2.column_dimensions[get_column_letter(col)].width = ancho

    wb.save(destino)
    logger.info(f"Excel generado en {destino}")
    return destino
